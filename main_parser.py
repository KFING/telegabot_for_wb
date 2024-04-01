import os
from random import choice
from typing import NamedTuple, Dict, Any, Set
import time
import logging
import telepot
import requests
from bs4 import BeautifulSoup
from config import desktop_agents_, list_time_work
import openpyxl


py_logger = logging.getLogger("game_stop")
py_logger.setLevel(logging.INFO)
py_handler = logging.FileHandler(f"{__name__}.log", mode='w')
py_formatter = logging.Formatter("%(name)s %(asctime)s %(levelname)s %(message)s")
py_handler.setFormatter(py_formatter)
py_logger.addHandler(py_handler)

CWD = os.getcwd()
RESULTS_DIR = os.path.join(CWD, 'data')

list_time = list_time_work
desktop_agents = desktop_agents_

def random_headers():
    return {'User-Agent': choice(desktop_agents),'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}

class ParsingError(Exception):
    def __init__(self, message: str, exception: Exception) -> None:
        self.exception = exception
        self.message = message

class RequestError(Exception):
    def __init__(self, message: str, exception: Exception) -> None:
        self.exception = exception
        self.message = message
class ArticleInfo(NamedTuple):
    info_text: str
    title: str
    sku: int
    link: str
    root: str
    link_sku: str
    grade: str
    text: str
    reting: str

    def to_dict(self) -> Dict[str, Any]:
        dct = self._asdict()
        return dct

def get_link(sku:str) -> str:
    return f'https://www.wildberries.ru/catalog/{sku}/feedbacks'



def parse_article__stop_game(url: str, sku: str) -> ArticleInfo:
    try:
        response = requests.get(url, headers=random_headers())
        response.raise_for_status()
        html = response.content.decode('utf-8')
    except Exception as e:
        raise RequestError('wrong request', e)

    try:
        text_c = ""
        soup = BeautifulSoup(html, 'html.parser')
        title_pub = soup.find('h1').text
        for content in soup.find("div", class_="_content_1gk4z_10").find_all("p",
                                                                             class_="_text_1gk4z_108 _text-width_1gk4z_108"):
            text_c += " " + content.text
        content_count = 250
        if(len(text_c) > 250):
            content_count = 200
            while (text_c[content_count] != '.') and (content_count < 350):
                content_count += 1
            if content_count > 348:
                content_count = 200
                while (text_c[content_count] != '.') and (content_count > 120):
                    content_count -= 1
        return ArticleInfo(
            title=title_pub,
            text=f'{text_c[:content_count]}. '
        )
    except Exception as e:
        raise ParsingError('wrong content of page', e)
def get_html(url:str) -> str:
    try:
        response = requests.get(url, headers=random_headers())
        response.raise_for_status()
        return response.content.decode('utf-8')
    except Exception as e:
        raise RequestError('wrong request', e)
def get_root(html: str) -> str:
    try:
        text_c = ""
        soup = BeautifulSoup(html, 'html.parser')
        print(html)
        tmp = soup.find('a')
        if (tmp.get('data-wba-header-name') == 'Login'):
            link_tmp = tmp.get('href')
            root = link[link.find('3D') : link[link.find('3D') : ].find('%')]
            print(root)
            return root
    except Exception as e:
        raise ParsingError('wrong content of page', e)
if __name__ == '__main__':
    wookbook = openpyxl.load_workbook("SKUInfo.xlsx")
    worksheet = wookbook.active
    #for i in range(0, worksheet.max_row):
    #    for col in worksheet.iter_cols(1, worksheet.max_column):
    #        sku = str(col[i].value)
    #        link = get_link(sku)
    sku = '39374520'
    link = 'https://www.wildberries.ru/catalog/39374520/feedbacks'
    html = get_html(link)
    print()
    root = get_root(html)
    link_sku = f'https://feedbacks2.wb.ru/feedbacks/v1/{root}'
    #article_info = parse_article__stop_game(link, sku)
    #print(article_info)
    #    print('')
